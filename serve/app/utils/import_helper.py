"""Excel 导入助手 — 模板下载 / 解析 / 逐行校验入库.

导入流程:
  1. 模板下载: 复用 logic.export_header_map() 做表头（中文标签）, 填数后上传
  2. 解析: 表头标签 → 字段名 (反向映射 export_header_map)
  3. 入库: 每行独立事务 (logic.create 自带 commit), 一行失败不影响其他行

安全:
  - module 只能解析 app.logics.{module}, 防任意类实例化
"""
import io

from openpyxl import Workbook, load_workbook

from app.logics.base import BaseLogic, BizError


def resolve_logic_module(module_name: str) -> BaseLogic:
    """从 app.logics.{module_name} 解析出 Logic 实例."""
    import importlib
    import inspect

    if not module_name or not module_name.replace("_", "").isalnum():
        raise BizError("模块名不合法")
    try:
        mod = importlib.import_module(f"app.logics.{module_name}")
    except Exception:
        raise BizError(f"逻辑模块不存在: {module_name}")

    logic_classes = [
        obj for _name, obj in inspect.getmembers(mod, inspect.isclass)
        if issubclass(obj, BaseLogic) and obj is not BaseLogic
    ]
    if not logic_classes:
        raise BizError(f"{module_name} 未找到 Logic 类")

    # 优先取与模块同名的 Logic（如 dict → DictLogic, 而非 DictItemLogic）
    expected = f"{module_name}logic"
    for cls in logic_classes:
        if cls.__name__.lower() == expected:
            return cls()
    return logic_classes[0]()


def build_template_bytes(logic: BaseLogic) -> bytes:
    """生成导入模板 XLSX（表头 = export_header_map 的中文标签）."""
    header_map = logic.export_header_map()
    if not header_map:
        raise BizError("该模块不支持导入（未配置导出表头）")

    wb = Workbook()
    ws = wb.active
    ws.title = "导入数据"
    labels = list(header_map.values())
    ws.append(labels)
    # 说明行: 字段名 (辅助用户理解, 导入时忽略)
    ws.append([f"字段名: {f}" for f in header_map.keys()])

    # 列宽
    for i in range(1, len(labels) + 1):
        ws.column_dimensions[chr(64 + i)].width = 20

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def parse_rows(logic: BaseLogic, content: bytes) -> list[dict]:
    """解析上传的 XLSX → 行 dict 列表.

    首行为表头（中文标签 → 字段名映射），其后为数据行。空行跳过。
    """
    header_map = logic.export_header_map()
    if not header_map:
        raise BizError("该模块不支持导入")

    label_to_field = {label: field for field, label in header_map.items()}

    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)

    header_row = next(rows, None)
    if not header_row:
        raise BizError("模板为空")

    # 首行: 若含「字段名:」前缀说明是注释行, 跳过长这样的模板结构
    headers = [str(h).strip() if h is not None else "" for h in header_row]
    # 若第二行是注释行则自动跳过
    second = next(rows, None)
    if second and any(str(c or "").startswith("字段名:") for c in second):
        second = next(rows, None)

    parsed = []
    row_iter = [second] if second is not None else []
    row_iter.extend(rows)
    for row in row_iter:
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue
        item = {}
        for idx, header in enumerate(headers):
            field = label_to_field.get(header)
            if field is None:
                continue
            value = row[idx] if idx < len(row) else None
            # None / 空字符串 不写入 (让模型默认值生效)
            if value is not None and str(value).strip() != "":
                item[field] = value
        if item:
            parsed.append(item)
    return parsed


async def import_rows(db, logic: BaseLogic, rows: list[dict]) -> dict:
    """逐行校验 + 入库, 返回 {imported, failed, errors:[{row, error}]}.

    每行独立事务: 单行失败不影响其他行 (logic.create 内部自带 commit/rollback)。
    """
    imported = 0
    failed = 0
    errors = []

    for idx, row in enumerate(rows, start=3):  # 模板第 3 行起为数据 (1 表头 2 注释)
        try:
            await logic.create(db, row)
            imported += 1
        except Exception as e:
            failed += 1
            errors.append({"row": idx, "error": str(e)})

    return {"imported": imported, "failed": failed, "errors": errors}
