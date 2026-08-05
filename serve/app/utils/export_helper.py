"""
数据导出工具

功能：
    - openpyxl write_only 模式（常量内存，支持大数据量）
    - 游标分块查询（id < lastRowId，每批 5000 行）
    - Redis 实时进度更新
    - 超过 MAX_ROWS 自动拆分多 XLSX 打包 ZIP

用法（由 BaseLogic.do_export 调用，业务代码不直接使用）：
    helper = ExportHelper(model=Order, fields=[...], header_map={...}, file_key="abc")
    path = await helper.export(db, filters, row_formatter)
"""

import os
import zipfile
import hashlib
import logging
import tempfile
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from sqlalchemy import select, func, asc
from sqlalchemy.ext.asyncio import AsyncSession
from app.utils.query import apply_filters
from app.services.redis import get_redis

logger = logging.getLogger("export")

# 默认值
CHUNK_SIZE = 5000       # 每次 DB 查询行数
MAX_ROWS = 200000       # 单文件最大行数
EXPORT_DIR = os.path.join(tempfile.gettempdir(), "base_exports")
PROGRESS_TTL = 3600     # 进度 key 过期时间


def generate_export_key(user_id: int) -> str:
    """生成导出 key：md5(user_id)[:12]_timestamp"""
    uid_hash = hashlib.md5(str(user_id).encode()).hexdigest()[:12]
    return f"{uid_hash}_{int(datetime.now().timestamp())}"


def validate_export_key(key: str, user_id: int) -> bool:
    """校验导出 key 归属"""
    parts = key.split("_", 1)
    if len(parts) != 2:
        return False
    expected = hashlib.md5(str(user_id).encode()).hexdigest()[:12]
    return parts[0] == expected


class ExportHelper:
    def __init__(
        self,
        model,
        fields: list[str],
        header_map: dict[str, str],
        file_key: str,
        allowed_filters: list[str] | None = None,
        max_rows: int = MAX_ROWS,
        chunk_size: int = CHUNK_SIZE,
    ):
        self.model = model
        self.fields = fields
        self.header_map = header_map
        self.file_key = file_key
        self.allowed_filters = allowed_filters
        self.max_rows = max_rows
        self.chunk_size = chunk_size
        self.total_rows = 0
        self.written_rows = 0

    async def export(
        self,
        db: AsyncSession,
        filters: dict = None,
        row_formatter=None,
        format_context: dict = None,
    ) -> str:
        """
        执行导出

        :param db:             数据库 Session
        :param filters:        过滤条件（QueryHelper 格式）
        :param row_formatter:  行格式化函数 (row_dict, context) -> row_dict
        :param format_context: 格式化上下文（传给 row_formatter）
        :return:               生成的文件路径
        """
        # 确保导出目录存在
        os.makedirs(EXPORT_DIR, exist_ok=True)

        # 构建查询条件
        conditions = apply_filters(self.model, filters or {}, self.allowed_filters)

        # 统计总行数
        count_stmt = select(func.count()).select_from(self.model).where(*conditions)
        self.total_rows = (await db.execute(count_stmt)).scalar_one()

        if self.total_rows == 0:
            # 空数据也生成一个只有表头的文件（流式风格，不累积 list）
            path = self._make_xlsx_path()
            wb = Workbook(write_only=True)
            ws = wb.create_sheet("Sheet1")
            ws.append(list(self.header_map.values()))
            wb.save(path)
            await self._set_progress(100)
            await self._set_file_path(path)
            return path

        r = await get_redis()
        await r.set(f"export@{self.file_key}", "0", ex=PROGRESS_TTL)

        if self.total_rows <= self.max_rows:
            path = await self._export_single(db, conditions, row_formatter, format_context)
        else:
            path = await self._export_zip(db, conditions, row_formatter, format_context)

        await self._set_progress(100)
        await self._set_file_path(path)
        return path

    # ==================== 单文件导出 ====================

    async def _export_single(self, db, conditions, row_formatter, format_context) -> str:
        """单文件导出：逐 chunk 写入（常量内存，不累积全量行）"""
        path = self._make_xlsx_path()
        wb = Workbook(write_only=True)
        ws = wb.create_sheet("Sheet1")

        headers = list(self.header_map.values())
        fields = list(self.header_map.keys())
        ws.append(headers)

        last_id = 0
        while True:
            chunk = await self._fetch_chunk_batch(db, conditions, last_id, self.chunk_size)
            if not chunk:
                break
            last_id = chunk[-1].get("id", 0)

            for row in chunk:
                if row_formatter:
                    row = row_formatter(row, format_context)
                ws.append([row.get(f, "") for f in fields])
                self.written_rows += 1

            # 进度（查询+写入阶段占 0~80%）
            progress = min(80, int(self.written_rows / self.total_rows * 80))
            await self._set_progress(progress)

        wb.save(path)
        logger.info(f"XLSX saved: {path} ({self.written_rows} rows)")
        return path

    # ==================== 多文件 ZIP 导出 ====================

    async def _export_zip(self, db, conditions, row_formatter, format_context) -> str:
        xlsx_files = []
        last_id = 0
        file_index = 0

        while self.written_rows < self.total_rows:
            file_index += 1
            xlsx_path = self._make_xlsx_path(suffix=f"_part{file_index}")
            rows = await self._fetch_chunk_batch(db, conditions, last_id, self.max_rows)

            if not rows:
                break

            last_id = rows[-1].get("id", 0)
            self._write_xlsx(xlsx_path, rows, row_formatter, format_context)
            xlsx_files.append(xlsx_path)

        # 打包 ZIP
        zip_path = os.path.join(EXPORT_DIR, f"{self.file_key}.zip")
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for f in xlsx_files:
                zf.write(f, os.path.basename(f))
                os.remove(f)  # 清理临时 xlsx

        return zip_path

    # ==================== 数据查询 ====================

    async def _fetch_chunk_batch(self, db, conditions, last_id: int, limit: int) -> list[dict]:
        """单次分块查询（整行读取，从 Model 实例取字段）"""
        pk_col = getattr(self.model, "id", None)

        stmt = select(self.model).where(*conditions)
        if last_id > 0 and pk_col is not None:
            stmt = stmt.where(pk_col > last_id)
        if pk_col is not None:
            stmt = stmt.order_by(asc(pk_col))
        stmt = stmt.limit(limit)

        result = await db.execute(stmt)
        rows = []
        for record in result.scalars().all():
            row_dict = {}
            for f in self.fields:
                val = getattr(record, f, "")
                if hasattr(val, "isoformat"):
                    val = val.isoformat()
                row_dict[f] = val if val is not None else ""
            # 确保 id 在结果中（游标用）
            if "id" not in row_dict and hasattr(record, "id"):
                row_dict["id"] = record.id
            rows.append(row_dict)

        return rows

    # ==================== XLSX 写入 ====================

    def _write_xlsx(self, path: str, rows: list[dict], row_formatter=None, format_context=None):
        """用 openpyxl write_only 模式写入 XLSX"""
        wb = Workbook(write_only=True)
        ws = wb.create_sheet("Sheet1")

        # 表头
        headers = list(self.header_map.values())
        fields = list(self.header_map.keys())
        ws.append(headers)

        # 数据行
        for row in rows:
            if row_formatter:
                row = row_formatter(row, format_context)

            ws.append([row.get(f, "") for f in fields])
            self.written_rows += 1

        wb.save(path)
        logger.info(f"XLSX saved: {path} ({self.written_rows} rows)")

    # ==================== 辅助 ====================

    def _make_xlsx_path(self, suffix: str = "") -> str:
        return os.path.join(EXPORT_DIR, f"{self.file_key}{suffix}.xlsx")

    async def _set_progress(self, progress: int):
        r = await get_redis()
        await r.set(f"export@{self.file_key}", str(progress), ex=PROGRESS_TTL)

    async def _set_file_path(self, path: str):
        r = await get_redis()
        await r.set(f"export_file@{self.file_key}", path, ex=PROGRESS_TTL)
